#!/usr/bin/env python3
"""
공장 DB 구축 + 이메일 보강 파이프라인
=====================================
데이터 소스 (전국)
  - data.go.kr 15105482  : 한국산업단지공단_전국등록공장현황 (XLSX 파일)
    https://www.data.go.kr/data/15105482/fileData.do
  - data.go.kr 15057023  : 경기도_공장등록 현황 (Open API)
    https://www.data.go.kr/data/15057023/openapi.do

출력
  - output/factory_db.xlsx              : 시트 2개 (KICOX_전국, 경기도)
  - output/factory_db_enriched_<시트>.xlsx : 이메일 + 출처 URL 보강 결과

원칙
  - 이메일은 절대 추정/생성하지 않음. 실제 페이지에서 추출된 것만 기록.
  - 모든 이메일 행에 발견 위치(URL), 추출 방법, 시각을 같이 저장.

사전 준비
  1) data.go.kr 회원가입 -> 15057023 활용신청 -> serviceKey 발급
  2) 15105482 파일은 페이지에서 직접 다운로드 -> ./input/ 폴더에 그대로 저장
  3) Naver Developers (https://developers.naver.com) 검색 API 등록
       -> Client ID / Client Secret (무료, 일 25,000회)
  4) (선택) Google Custom Search API key + cx, 또는 SerpAPI key
  5) pip install pandas openpyxl requests beautifulsoup4 lxml tqdm tenacity python-dotenv

.env 예시 (스크립트와 같은 폴더에 두기)
  DATA_GO_KR_SERVICE_KEY=...
  GG_API_URL=https://api.odcloud.kr/api/15057023/v1/uddi:xxxxxxxx-...
  NAVER_CLIENT_ID=...
  NAVER_CLIENT_SECRET=...
  GOOGLE_CSE_API_KEY=...     # 선택
  GOOGLE_CSE_CX=...          # 선택
  SERPAPI_KEY=...            # 선택

실행
  python factory_db_pipeline.py build_db
  python factory_db_pipeline.py enrich --sheet 경기도 --limit 50      # 테스트
  python factory_db_pipeline.py enrich --sheet KICOX_전국             # 본작업
  python factory_db_pipeline.py enrich --sheet KICOX_전국 --resume    # 중단 지점부터 재개
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Set, Tuple
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential
from tqdm import tqdm

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ============================================================
# 설정
# ============================================================
INPUT_DIR  = Path("./input");  INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR = Path("./output"); OUTPUT_DIR.mkdir(exist_ok=True)

KICOX_FILE_GLOB = "한국산업단지공단_전국등록공장현황*.*"

DATA_GO_KR_KEY = os.getenv("DATA_GO_KR_SERVICE_KEY", "")
GG_API_URL     = os.getenv("GG_API_URL", "")

NAVER_CLIENT_ID     = os.getenv("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET", "")
GOOGLE_CSE_API_KEY  = os.getenv("GOOGLE_CSE_API_KEY", "")
GOOGLE_CSE_CX       = os.getenv("GOOGLE_CSE_CX", "")

SOURCE_URL_KICOX = "https://www.data.go.kr/data/15105482/fileData.do"
SOURCE_URL_GG    = "https://www.data.go.kr/data/15057023/openapi.do"

CRAWL_DELAY_SEC = 1.5
TIMEOUT_SEC     = 12
USER_AGENT      = "Mozilla/5.0 (compatible; AblearnBD-Research/1.0)"

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
EMAIL_BLOCKLIST_DOMAINS = {
    "example.com", "domain.com", "yourdomain.com", "test.com", "email.com",
    "sentry.io", "wixpress.com", "godo.co.kr", "cafe24.com",
    "sentry-next.wixpress.com", "saramin.co.kr", "happycampus.com",
    "jobkorea.co.kr", "albamon.com", "incruit.com", "work.go.kr",
    "kreditjob.com", "catch.co.kr", "114.co.kr"
}
EMAIL_BLOCKLIST_LOCAL = {"noreply", "no-reply", "donotreply", "do-not-reply", "example", "test", "email", "yourname", "yourid", "abc", "xxx"}

URL_BLOCKLIST = [
    "saramin.co.kr", "jobkorea.co.kr", "albamon.com", "incruit.com", "work.go.kr", 
    "catch.co.kr", "happycampus.com", "114.co.kr", "bizno.net", "kreditjob.com",
    "nicebizinfo.com", "rocketpunch.com", "wanted.co.kr", "jobplanet.co.kr"
]

SAVE_EVERY_N_ROWS = 200  # 주기적 체크포인트

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("factory_db")

# 서드파티 라이브러리들의 불필요한 내부 로그(INFO) 숨기기
for logger_name in ["ddgs", "duckduckgo_search", "urllib3", "httpx", "httpcore", "curl_cffi"]:
    logging.getLogger(logger_name).setLevel(logging.WARNING)


# ============================================================
# Phase 1-A. KICOX 전국 파일 로드
# ============================================================
def load_kicox() -> pd.DataFrame:
    files = list(INPUT_DIR.glob(KICOX_FILE_GLOB))
    # csv와 xlsx만 필터링
    valid_files = [f for f in files if f.suffix.lower() in [".csv", ".xlsx"]]
    if not valid_files:
        raise FileNotFoundError(
            f"{INPUT_DIR}/ 안에 KICOX 파일(.csv 또는 .xlsx)이 없음. "
            f"data.go.kr/data/15105482 에서 다운로드 후 그대로 저장."
        )
    path = sorted(valid_files)[-1]
    log.info("KICOX 파일 로드: %s", path)
    
    if path.suffix.lower() == ".csv":
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8").fillna("")
        except UnicodeDecodeError:
            df = pd.read_csv(path, dtype=str, encoding="cp949", encoding_errors="replace").fillna("")
    else:
        df = pd.read_excel(path, dtype=str).fillna("")

    col_map = {}
    for c in df.columns:
        cs = c.strip()
        if "회사" in cs or "업체" in cs:    col_map[c] = "회사명"
        elif "산업단지" in cs:               col_map[c] = "산업단지명"
        elif "생산" in cs or "품목" in cs:   col_map[c] = "생산품"
        elif "주소" in cs:                   col_map[c] = "공장주소"
    df = df.rename(columns=col_map)

    keep = [c for c in ["회사명", "산업단지명", "생산품", "공장주소"] if c in df.columns]
    df = df[keep].copy()
    df["출처_데이터셋"] = "한국산업단지공단_전국등록공장현황(15105482)"
    df["출처_URL"]      = SOURCE_URL_KICOX
    df = df.drop_duplicates(subset=keep).reset_index(drop=True)
    log.info("KICOX 행 수: %d", len(df))
    return df


# ============================================================
# Phase 1-B. 경기도 Open API
# ============================================================
@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, max=20),
    retry=retry_if_exception_type(requests.RequestException),
)
def _gg_call(page: int, per_page: int) -> dict:
    if not (DATA_GO_KR_KEY and GG_API_URL):
        raise RuntimeError(
            "DATA_GO_KR_SERVICE_KEY 와 GG_API_URL 환경변수 필요. "
            "data.go.kr/data/15057023 활용신청 후 .env 설정."
        )
    if "openapi.gg.go.kr" in GG_API_URL:
        params = {
            "pIndex": page,
            "pSize": per_page,
            "KEY": DATA_GO_KR_KEY,
            "Type": "json",
        }
    else:
        params = {
            "page": page,
            "perPage": per_page,
            "serviceKey": DATA_GO_KR_KEY,
            "returnType": "JSON",
        }
    r = requests.get(GG_API_URL, params=params, timeout=TIMEOUT_SEC,
                     headers={"User-Agent": USER_AGENT})
    r.raise_for_status()
    try:
        return r.json()
    except json.JSONDecodeError:
        # XML 폴백 (data.go.kr 용)
        soup = BeautifulSoup(r.content, "lxml-xml")
        items = []
        for it in soup.find_all("item"):
            items.append({c.name: (c.text or "") for c in it.find_all() if c.name})
        total = soup.find("totalCount")
        return {
            "data": items,
            "totalCount": int(total.text) if total and total.text else len(items),
        }


def load_gyeonggi() -> pd.DataFrame:
    # 1. 로컬에 저장된 엑셀 파일이 있으면 우선 로드 (시간 절약)
    local_path = INPUT_DIR / "경기도_공장등록현황.xlsx"
    if local_path.exists():
        log.info(f"로컬 파일 발견, API 대신 캐시 로드: {local_path}")
        df = pd.read_excel(local_path, dtype=str).fillna("")
        # 컬럼 매핑 (API 원본 컬럼명 -> 통일된 컬럼명)
        col_map = {}
        for c in df.columns:
            cs = c.strip()
            if cs in ("CMPNY_NM", "COMPNY_GRP_NM", "업체명", "회사명", "기업체명"):     col_map[c] = "업체명"
            elif "주소" in cs or "ADDR" in cs.upper():                col_map[c] = "공장주소"
            elif "연락" in cs or "TEL" in cs.upper() or "PHONE" in cs.upper(): col_map[c] = "대표연락처"
            elif "생산" in cs or "PROD" in cs.upper() or "PRDT" in cs.upper(): col_map[c] = "주요생산품"
            elif "업종" in cs or "INDU" in cs.upper():              col_map[c] = "업종명"
        df = df.rename(columns=col_map)
        
        expected = [c for c in ["업체명", "공장주소", "대표연락처", "주요생산품", "업종명"] if c in df.columns]
        df = df[expected].copy() if expected else df
        df["출처_데이터셋"] = "경기도_공장등록 현황(15057023)"
        df["출처_URL"]      = SOURCE_URL_GG
        log.info("경기도 행 수: %d", len(df))
        return df

    # 2. 로컬 파일이 없으면 API 호출
    if not (DATA_GO_KR_KEY and GG_API_URL):
        log.warning("경기도 API 키 미설정 -> 빈 시트로 출력")
        return pd.DataFrame(columns=[
            "업체명", "공장주소", "대표연락처", "주요생산품", "업종명",
            "출처_데이터셋", "출처_URL",
        ])

    log.info("경기도 API 페이징 시작")
    rows: List[dict] = []
    page = 1
    per_page = 1000
    total: Optional[int] = None
    while True:
        try:
            payload = _gg_call(page, per_page)
        except Exception as e:
            log.error("경기도 API page=%d 실패: %s", page, e)
            break
        data = (
            payload.get("data")
            or payload.get("response", {}).get("body", {}).get("items")
        )
        # 경기데이터드림 (FACTRYREGISTTM) 지원
        if not data and "FACTRYREGISTTM" in payload:
            for item in payload["FACTRYREGISTTM"]:
                if "row" in item:
                    data = item["row"]
                    break
        data = data or []
        
        if not data:
            break
        rows.extend(data)
        if total is None:
            total = (
                payload.get("totalCount")
                or payload.get("response", {}).get("body", {}).get("totalCount")
            )
            if not total and "FACTRYREGISTTM" in payload:
                for item in payload["FACTRYREGISTTM"]:
                    if "head" in item:
                        total = item["head"][0].get("list_total_count")
                        break
            if total:
                log.info("총 데이터 건수: %s", total)
        log.info("page=%d 누적 %d행", page, len(rows))
        if total and len(rows) >= int(total):
            break
        page += 1
        time.sleep(0.5)

    df = pd.DataFrame(rows).fillna("")
    col_map = {}
    for c in df.columns:
        cs = c.strip()
        if cs in ("CMPNY_NM", "COMPNY_GRP_NM", "업체명", "회사명", "기업체명"):     col_map[c] = "업체명"
        elif "주소" in cs or "ADDR" in cs.upper():                col_map[c] = "공장주소"
        elif "연락" in cs or "TEL" in cs.upper() or "PHONE" in cs.upper(): col_map[c] = "대표연락처"
        elif "생산" in cs or "PROD" in cs.upper() or "PRDT" in cs.upper(): col_map[c] = "주요생산품"
        elif "업종" in cs or "INDU" in cs.upper():              col_map[c] = "업종명"
    df = df.rename(columns=col_map)

    expected = [c for c in ["업체명", "공장주소", "대표연락처", "주요생산품", "업종명"] if c in df.columns]
    df = df[expected].copy() if expected else df
    df["출처_데이터셋"] = "경기도_공장등록 현황(15057023)"
    df["출처_URL"]      = SOURCE_URL_GG
    log.info("경기도 행 수: %d", len(df))
    return df


# ============================================================
# Phase 1-C. Excel 출력 (시트 2개, 그레이스케일)
# ============================================================
def save_db_excel(kicox: pd.DataFrame, gg: pd.DataFrame, out_path: Path) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        kicox.to_excel(writer, sheet_name="KICOX_전국", index=False)
        gg.to_excel(writer, sheet_name="경기도", index=False)
        wb = writer.book
        for sheet_name, df in [("KICOX_전국", kicox), ("경기도", gg)]:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9D9D9")  # 그레이스케일
                cell.alignment = Alignment(horizontal="left", vertical="center")
            for i, col in enumerate(df.columns, start=1):
                sample = df[col].astype(str).head(200)
                width = max([len(str(col))] + [min(len(v), 60) for v in sample]) + 2
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(12, width), 60)
            ws.freeze_panes = "A2"
    log.info("저장 완료: %s", out_path)


def cmd_build_db(_args) -> None:
    kicox = load_kicox()
    gg    = load_gyeonggi()
    save_db_excel(kicox, gg, OUTPUT_DIR / "factory_db.xlsx")


# ============================================================
# Phase 2. 이메일 다중 소스 크롤링
# ============================================================
@dataclass
class EmailHit:
    email: str
    source_url: str
    method: str  # naver_webkr, naver_local, google_cse, serpapi, homepage
    found_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))

    def is_valid(self) -> bool:
        em = self.email.lower().strip().rstrip(".,;:")
        local, _, domain = em.partition("@")
        if not domain or "." not in domain:                 return False
        if domain in EMAIL_BLOCKLIST_DOMAINS:                return False
        if local in EMAIL_BLOCKLIST_LOCAL:                   return False
        if domain.endswith((".png", ".jpg", ".gif", ".pdf", ".zip", ".webp")): return False
        if len(em) > 80 or len(local) < 1:                   return False
        return True


def _clean(em: str) -> str:
    return em.strip().rstrip(".,;:'\"")


# --- Naver Open API (webkr / local) ---
@retry(stop=stop_after_attempt(2), wait=wait_exponential(multiplier=1, max=5))
def naver_search(query: str, kind: str = "webkr", display: int = 10) -> List[dict]:
    if not (NAVER_CLIENT_ID and NAVER_CLIENT_SECRET):
        return []
    url = f"https://openapi.naver.com/v1/search/{kind}.json"
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    r = requests.get(url, headers=headers,
                     params={"query": query, "display": display},
                     timeout=TIMEOUT_SEC)
    if r.status_code != 200:
        log.debug("Naver %s 실패 %d %s", kind, r.status_code, r.text[:100])
        return []
    return r.json().get("items", [])


# --- Google Custom Search ---
def google_cse(query: str, num: int = 10) -> List[dict]:
    if not (GOOGLE_CSE_API_KEY and GOOGLE_CSE_CX):
        return []
    try:
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": GOOGLE_CSE_API_KEY, "cx": GOOGLE_CSE_CX, "q": query, "num": min(num, 10)},
            timeout=TIMEOUT_SEC,
        )
        if r.status_code != 200:
            return []
        return r.json().get("items", [])
    except requests.RequestException:
        return []


# --- DuckDuckGo Search (무료 대안) ---
from ddgs import DDGS

def duckduckgo_search(query: str, num: int = 5) -> List[dict]:
    try:
        with DDGS() as ddgs:
            # DuckDuckGo 라이브러리는 max_results 개수만큼 결과를 반환
            results = ddgs.text(query, max_results=num, region="kr-kr")
            if not results:
                return []
            return [{"link": r.get("href", ""), "title": r.get("title", ""), "snippet": r.get("body", "")} for r in results]
    except Exception as e:
        log.debug("DuckDuckGo 검색 실패: %s", e)
        return []


# --- 페이지 fetch + 이메일 추출 ---
@retry(stop=stop_after_attempt(2), wait=wait_exponential(multiplier=1, max=5))
def fetch_html(url: str) -> str:
    r = requests.get(url, timeout=TIMEOUT_SEC, headers={"User-Agent": USER_AGENT})
    r.raise_for_status()
    # encoding 추정
    if r.encoding is None or r.encoding.lower() == "iso-8859-1":
        r.encoding = r.apparent_encoding
    return r.text


def extract_emails_from_text(text: str) -> Set[str]:
    raw = set(EMAIL_REGEX.findall(text or ""))
    out: Set[str] = set()
    for em in raw:
        em = _clean(em)
        local, _, domain = em.lower().partition("@")
        if not domain or "." not in domain:                                continue
        if domain in EMAIL_BLOCKLIST_DOMAINS:                              continue
        if local in EMAIL_BLOCKLIST_LOCAL:                                 continue
        if domain.endswith((".png", ".jpg", ".gif", ".pdf", ".webp")):     continue
        if "@2x" in em or "@3x" in em:                                     continue  # 이미지 retina
        out.add(em)
    return out


def find_contact_pages(soup: BeautifulSoup, base_url: str) -> List[str]:
    keywords = ("contact", "about", "company", "introduction",
                "연락", "문의", "회사소개", "오시는길", "개요")
    out: List[str] = []
    seen: Set[str] = set()
    base_host = urlparse(base_url).netloc
    for a in soup.find_all("a", href=True):
        href = a["href"]
        txt = (a.get_text() or "").lower()
        if any(k in href.lower() or k in txt for k in keywords):
            full = urljoin(base_url, href)
            if urlparse(full).netloc == base_host and full not in seen:
                seen.add(full)
                out.append(full)
        if len(out) >= 5:
            break
    return out


def scrape_emails_from_url(url: str) -> List[EmailHit]:
    hits: List[EmailHit] = []
    try:
        html = fetch_html(url)
    except Exception as e:
        log.debug("fetch 실패 %s: %s", url, e)
        return hits
    soup = BeautifulSoup(html, "lxml")
    for em in extract_emails_from_text(html):
        hits.append(EmailHit(email=em, source_url=url, method="homepage"))
    for sub in find_contact_pages(soup, url):
        time.sleep(CRAWL_DELAY_SEC)
        try:
            html2 = fetch_html(sub)
        except Exception:
            continue
        for em in extract_emails_from_text(html2):
            hits.append(EmailHit(email=em, source_url=sub, method="homepage"))
    return hits


def search_emails_for_company(name: str, address_hint: str = "") -> List[EmailHit]:
    """다중 소스로 후보 URL/스니펫 수집 -> 페이지 본문에서 이메일 추출."""
    hits: List[EmailHit] = []
    seen_emails: Set[str] = set()
    seen_urls:   Set[str] = set()

    queries = [
        f"{name} 이메일",
        f"{name} 문의",
        f"{name} 공식 홈페이지",
        f"{name} contact email",
    ]
    if address_hint:
        city = address_hint.split()[0] if address_hint.split() else ""
        if city:
            queries.append(f"{name} {city}")

    candidates: List[Tuple[str, str, str]] = []  # (url, snippet, method)

    for q in queries:
        tqdm.write(f"    - 🔍 [Naver 웹문서] 검색 중: '{q}'")
        for it in naver_search(q, "webkr", 10):
            url = it.get("link") or ""
            snip = (it.get("title", "") + " " + it.get("description", ""))
            if url and url not in seen_urls and not any(b in url for b in URL_BLOCKLIST):
                seen_urls.add(url); candidates.append((url, snip, "naver_webkr"))
        time.sleep(0.3)

        tqdm.write(f"    - 🔍 [Naver 지역정보] 검색 중: '{q}'")
        for it in naver_search(q, "local", 5):
            url = it.get("link") or ""
            snip = " ".join([it.get("title", ""), it.get("description", ""), it.get("address", "")])
            if url and url not in seen_urls and not any(b in url for b in URL_BLOCKLIST):
                seen_urls.add(url); candidates.append((url, snip, "naver_local"))
        time.sleep(0.3)

        tqdm.write(f"    - 🔍 [Google Custom Search] 검색 중: '{q}'")
        for it in google_cse(q, 10):
            url = it.get("link") or ""
            snip = (it.get("title", "") + " " + it.get("snippet", ""))
            if url and url not in seen_urls and not any(b in url for b in URL_BLOCKLIST):
                seen_urls.add(url); candidates.append((url, snip, "google_cse"))
        time.sleep(0.3)

        tqdm.write(f"    - 🔍 [DuckDuckGo(무료)] 검색 중: '{q}'")
        for it in duckduckgo_search(q, 5):
            url = it.get("link") or ""
            snip = (it.get("title", "") + " " + it.get("snippet", ""))
            if url and url not in seen_urls and not any(b in url for b in URL_BLOCKLIST):
                seen_urls.add(url); candidates.append((url, snip, "duckduckgo"))
        time.sleep(0.5)

    # 1) 검색결과 스니펫에서 직접 이메일 추출 (가장 빠른 hit)
    for url, snippet, method in candidates:
        for em in extract_emails_from_text(snippet):
            if em not in seen_emails:
                seen_emails.add(em)
                hits.append(EmailHit(email=em, source_url=url, method=f"{method}_snippet"))
                tqdm.write(f"      => [스니펫 발견] {em} ({method})")

    # 2) 후보 URL 직접 방문 (상위 N개만 - 시간 절약)
    to_visit = [u for u, _, _ in candidates[:8] if not any(u.lower().endswith(ext) for ext in (".pdf", ".jpg", ".png", ".gif", ".zip"))]
    if to_visit:
        tqdm.write(f"    - 홈페이지 방문 탐색 진행: {len(to_visit)}개 URL 접속 중...")
        
    for url in to_visit:
        time.sleep(CRAWL_DELAY_SEC)
        found_in_url = False
        for hit in scrape_emails_from_url(url):
            if hit.email not in seen_emails:
                seen_emails.add(hit.email)
                hits.append(hit)
                found_in_url = True
                tqdm.write(f"      => [홈페이지 발견] {hit.email} ({url})")
        if found_in_url:
            break # 한 URL에서 찾으면 다음 URL은 생략 (속도 우선)

    return [h for h in hits if h.is_valid()]


# ============================================================
# Phase 2 main: enrich
# ============================================================
def cmd_enrich(args) -> None:
    in_path = OUTPUT_DIR / "factory_db.xlsx"
    if not in_path.exists():
        raise FileNotFoundError("먼저 build_db 실행 필요.")

    sheet = args.sheet
    df = pd.read_excel(in_path, sheet_name=sheet, dtype=str).fillna("")
    name_col = "회사명" if "회사명" in df.columns else "업체명"
    addr_col = "공장주소" if "공장주소" in df.columns else None

    out_path = OUTPUT_DIR / f"factory_db_enriched_{sheet}.xlsx"

    # 재개 모드: 기존 결과의 행 인덱스 이후만 처리
    processed_idx: Set[int] = set()
    out_rows: List[dict] = []
    if args.resume and out_path.exists():
        prev = pd.read_excel(out_path, dtype=str).fillna("")
        if "_원본행번호" in prev.columns:
            processed_idx = set(prev["_원본행번호"].astype(int).tolist())
            out_rows = prev.to_dict(orient="records")
            log.info("재개: 기처리 %d행 스킵", len(processed_idx))

    # 청크(Chunk) 분할 적용
    start_i = args.start_idx if args.start_idx else 0
    end_i = args.end_idx if args.end_idx else len(df)
    
    # args.limit이 있으면 end_i를 start_i + limit 으로 덮어씀 (호환성 유지)
    if args.limit:
        end_i = start_i + args.limit

    df = df.iloc[start_i:end_i].copy()
    log.info(f"작업 범위: 인덱스 {start_i} ~ {end_i-1} (총 {len(df)}행)")

    n_with_email = 0
    save_counter = 0

    for i, row in tqdm(list(df.iterrows()), total=len(df), desc=f"이메일 보강 [{sheet}]"):
        if i in processed_idx:
            continue
        company = (row.get(name_col) or "").strip()
        addr    = (row.get(addr_col) or "").strip() if addr_col else ""
        base    = row.to_dict()
        base["_원본행번호"] = i

        tqdm.write(f"\n[작업 중] 회사명: {company} (인덱스: {i})")

        if not company:
            out_rows.append({**base, "이메일": "", "이메일_출처_URL": "",
                             "이메일_방법": "", "이메일_조회시각": ""})
            tqdm.write("    - 스킵: 회사명 없음")
        else:
            try:
                hits = search_emails_for_company(company, addr)
            except Exception as e:
                log.warning("크롤링 실패 [%s]: %s", company, e)
                hits = []

            if not hits:
                out_rows.append({**base, "이메일": "", "이메일_출처_URL": "",
                                 "이메일_방법": "", "이메일_조회시각": ""})
            else:
                # 한 회사 여러 이메일은 행 분할(추적성 우선)
                for h in hits:
                    out_rows.append({**base,
                                     "이메일": h.email,
                                     "이메일_출처_URL": h.source_url,
                                     "이메일_방법": h.method,
                                     "이메일_조회시각": h.found_at})
                n_with_email += 1

        save_counter += 1
        if save_counter % SAVE_EVERY_N_ROWS == 0:
            pd.DataFrame(out_rows).to_excel(out_path, index=False)
            log.info("체크포인트 저장 (처리 %d / 이메일 보유 %d)", save_counter, n_with_email)

    enriched = pd.DataFrame(out_rows)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        enriched.to_excel(writer, sheet_name=sheet, index=False)

    have_mail = enriched["이메일"].astype(bool).sum() if "이메일" in enriched.columns else 0
    log.info("최종 저장: %s (총 %d행, 이메일 발견 %d행)", out_path, len(enriched), have_mail)


# ============================================================
# Entry point
# ============================================================
def main() -> None:
    parser = argparse.ArgumentParser(description="공장 DB + 이메일 보강 파이프라인")
    sub = parser.add_subparsers(dest="cmd", required=True)
    sub.add_parser("build_db", help="Phase 1: KICOX 파일 + 경기도 API -> Excel 시트 2개")
    e = sub.add_parser("enrich", help="Phase 2: 이메일 크롤링 + 출처 URL 기록")
    e.add_argument("--sheet",  required=True, choices=["KICOX_전국", "경기도"])
    e.add_argument("--limit",  type=int, default=0, help="시험용 상위 N개만")
    e.add_argument("--start-idx", type=int, default=0, help="시작할 행 인덱스 (0부터 시작)")
    e.add_argument("--end-idx", type=int, default=0, help="끝날 행 인덱스")
    e.add_argument("--resume", action="store_true", help="중단 지점부터 재개")
    args = parser.parse_args()
    if args.cmd == "build_db":
        cmd_build_db(args)
    elif args.cmd == "enrich":
        cmd_enrich(args)


if __name__ == "__main__":
    main()