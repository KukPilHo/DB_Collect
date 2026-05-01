"""
엑셀 엑스포트
==============
enrichment_state JOIN master → output/outbound_db_<날짜>.xlsx
시트 구성: 카테고리 4시트(best-pick) + 통합(best-pick) + 상세(전체 이메일) + summary
그레이스케일 스타일링 (기존 save_db_excel 이식).
"""

from datetime import date

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import (
    CATEGORY_QUOTA,
    MASTER_PATH,
    OUTPUT_DIR,
    log,
)
from src.enricher.best_pick import select_best_picks
from src.enricher.state import EnrichmentState, STATUS_COLLECTED


def _style_sheet(ws, df: pd.DataFrame) -> None:
    """그레이스케일 엑셀 스타일링: 헤더 #D9D9D9, bold, freeze A2."""
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 컬럼 너비 자동 조정
    for i, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).fillna("").head(200)
        width = max([len(str(col))] + [min(len(str(v)), 60) for v in sample]) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(12, width), 60)

    ws.freeze_panes = "A2"


def run_export() -> str:
    """enrichment_state + master → 최종 엑셀 생성."""
    if not MASTER_PATH.exists():
        raise FileNotFoundError("master.parquet 없음.")

    master = pd.read_parquet(MASTER_PATH)
    state = EnrichmentState()
    emails = state.get_all_emails()
    state.close()

    # 이메일 데이터를 DataFrame으로
    email_df = pd.DataFrame(emails)

    if email_df.empty:
        log.warning("enrichment_state에 데이터 없음. 빈 엑셀 생성.")
        email_df = pd.DataFrame(columns=["회사키", "이메일", "이메일_출처_URL", "이메일_방법", "상태"])

    # ── Best-Pick 선택 (회사당 최대 2건) ──────────────────────
    best_picks = select_best_picks(emails, max_picks=2)
    best_pick_df = pd.DataFrame(best_picks) if best_picks else pd.DataFrame(
        columns=["회사키", "이메일", "이메일_출처_URL", "이메일_방법", "상태", "best_pick_점수", "best_pick_순위"]
    )

    # ── 전체 collected 이메일 (상세 시트용) ────────────────────
    if not email_df.empty:
        all_collected = email_df[email_df["상태"] == STATUS_COLLECTED].copy()
    else:
        all_collected = pd.DataFrame(columns=["회사키", "이메일", "이메일_출처_URL", "이메일_방법", "상태"])

    # ── master와 JOIN ─────────────────────────────────────────
    def _merge_with_master(email_subset: pd.DataFrame) -> pd.DataFrame:
        """email subset을 master와 조인."""
        if email_subset.empty:
            return master.head(0)
        joined = email_subset.drop(columns=["회사명"], errors="ignore")
        return master.merge(joined, on="회사키", how="inner")

    merged_best = _merge_with_master(best_pick_df)
    merged_all = _merge_with_master(all_collected)

    # 출력 컬럼 선택 (조회시각을 P열에 배치)
    out_cols = [
        "회사명", "이메일", "이메일_출처_URL", "이메일_방법",
        "업종명_원본", "시도", "시군구", "도로명주소", "대표전화", "홈페이지",
        "종업원수", "자본금", "카테고리", "출처_데이터셋ID", "조회시각"
    ]
    out_cols = [c for c in out_cols if c in merged_best.columns]

    # 조회시각 기준으로 오름차순 정렬 (새로 수집된 데이터가 맨 밑으로 쌓이게 함)
    if "조회시각" in merged_best.columns:
        merged_best = merged_best.sort_values(by="조회시각", ascending=True)
    if "조회시각" in merged_all.columns:
        merged_all = merged_all.sort_values(by="조회시각", ascending=True)

    # 상세 시트 컬럼 (best_pick 점수 포함)
    detail_cols = out_cols.copy()
    if "best_pick_점수" in merged_best.columns:
        detail_extra = ["best_pick_점수", "best_pick_순위"]
        detail_extra = [c for c in detail_extra if c in merged_best.columns]
    else:
        detail_extra = []

    out_path = OUTPUT_DIR / f"outbound_db_{date.today().isoformat()}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── 카테고리별 시트 (Best-Pick만) ─────────────────────
        categories = list(CATEGORY_QUOTA.keys())
        summary_data = []

        for cat in categories:
            cat_df = merged_best[merged_best["카테고리"] == cat][out_cols].copy() if not merged_best.empty else pd.DataFrame(columns=out_cols)
            if not cat_df.empty:
                # 회사명 기준으로 고유 번호(dense rank) 부여
                cat_df.insert(0, "No.", pd.factorize(cat_df["회사명"])[0] + 1)
            else:
                cat_df.insert(0, "No.", [])
                
            sheet_name = cat[:31]  # 엑셀 시트명 31자 제한
            cat_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            _style_sheet(ws, cat_df)

            # summary 데이터
            quota = CATEGORY_QUOTA.get(cat, 0)
            unique_companies = cat_df["회사명"].nunique() if not cat_df.empty else 0
            summary_data.append({
                "카테고리": cat,
                "목표(quota)": quota,
                "Best-Pick 행수": len(cat_df),
                "고유 회사 수": unique_companies,
                "미달/초과": unique_companies - quota,
            })

        # ── 통합 시트 (Best-Pick) ─────────────────────────────
        if not merged_best.empty:
            merged_out = merged_best[out_cols].copy()
            merged_out.insert(0, "No.", pd.factorize(merged_out["회사명"])[0] + 1)
            merged_out.to_excel(writer, sheet_name="통합", index=False)
            ws = writer.book["통합"]
            _style_sheet(ws, merged_out)
        else:
            empty_df = pd.DataFrame(columns=["No."] + out_cols)
            empty_df.to_excel(writer, sheet_name="통합", index=False)

        # ── 상세 시트 (전체 collected 이메일) ─────────────────
        all_out_cols = [c for c in out_cols if c in merged_all.columns]
        if not merged_all.empty:
            detail_df = merged_all[all_out_cols].copy()
            detail_df.insert(0, "No.", pd.factorize(detail_df["회사명"])[0] + 1)
            detail_df.to_excel(writer, sheet_name="상세_전체이메일", index=False)
            ws = writer.book["상세_전체이메일"]
            _style_sheet(ws, detail_df)
        else:
            empty_detail = pd.DataFrame(columns=["No."] + all_out_cols)
            empty_detail.to_excel(writer, sheet_name="상세_전체이메일", index=False)

        # ── summary 시트 ──────────────────────────────────────
        summary_df = pd.DataFrame(summary_data)

        # hit rate 계산
        state2 = EnrichmentState()
        status_counts = state2.count_by_status()
        total_processed = sum(status_counts.values())
        collected_count = status_counts.get(STATUS_COLLECTED, 0)
        hit_rate = (collected_count / total_processed * 100) if total_processed > 0 else 0
        state2.close()

        hit_info = pd.DataFrame([{
            "전체 처리": total_processed,
            "이메일 발견": collected_count,
            "hit rate (%)": f"{hit_rate:.1f}",
        }])

        best_pick_info = pd.DataFrame([{
            "Best-Pick 총 행수": len(best_pick_df),
            "Best-Pick 고유 회사 수": best_pick_df["회사키"].nunique() if not best_pick_df.empty else 0,
            "전체 이메일 행수": len(all_collected),
        }])

        summary_df = pd.concat([summary_df, pd.DataFrame([{}]), hit_info, pd.DataFrame([{}]), best_pick_info], ignore_index=True)

        summary_df.to_excel(writer, sheet_name="summary", index=False)
        ws = writer.book["summary"]
        _style_sheet(ws, summary_df)

    log.info("엑셀 저장 완료: %s", out_path)
    log.info("  Best-Pick: %d행 (%d개 회사)", len(best_pick_df),
             best_pick_df["회사키"].nunique() if not best_pick_df.empty else 0)
    log.info("  상세(전체): %d행", len(all_collected))
    return str(out_path)
