"""
엑셀 기반 이메일 수집
=======================
사용자가 제공한 엑셀 파일의 회사 목록을 대상으로 이메일을 수집한다.
기존 7단계 파이프라인(download → ... → export)과 독립적으로 동작.

사용법:
  python -m src.cli enrich-excel --file 내회사목록.xlsx
  python -m src.cli enrich-excel --file 내회사목록.xlsx --skip-existing
  python -m src.cli enrich-excel --file 내회사목록.xlsx --limit 10
"""

import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm

from src.config import OUTPUT_DIR, log
from src.enricher.best_pick import select_best_picks
from src.enricher.search import search_emails_for_company, fast_search_single_email
from src.enricher.state import (
    EnrichmentState,
    STATUS_COLLECTED,
    STATUS_FAILED,
    STATUS_NOT_FOUND,
    STATUS_SKIPPED,
)

# ── 회사명 정규화 (중복 체크용) ──────────────────────────────
_CORP_PATTERNS = re.compile(
    r"주식회사|유한회사|사단법인|재단법인|합자회사|합명회사"
    r"|\(주\)|\(유\)|\(사\)|\(재\)|\(합\)"
    r"|\(주식회사\)|\(유한회사\)|\(사단법인\)|\(재단법인\)"
    r"|㈜|㈜"
)
_SPECIAL_CHARS = re.compile(r"[()（）「」『』\[\]<>《》\"'.,·•]")


def _normalize_name(name: str) -> str:
    """회사명 정규화: 법인격/특수문자/공백 제거 → lower."""
    if not name or not isinstance(name, str):
        return ""
    s = _CORP_PATTERNS.sub("", name)
    s = _SPECIAL_CHARS.sub("", s)
    s = re.sub(r"\s+", "", s)
    return s.lower()


# ── 컬럼명 자동 감지 ─────────────────────────────────────────

# 회사명 후보 컬럼명 (우선순위 순)
_NAME_COL_CANDIDATES = ["회사명", "업체명", "기업명", "기업체명", "상호", "상호명", "법인명", "사업체명", "수행기관명", "수행기관", "기관명", "company", "name"]

# 주소 후보 컬럼명
_ADDR_COL_CANDIDATES = ["주소", "공장주소", "도로명주소", "지번주소", "소재지", "address", "addr"]

# 홈페이지 후보 컬럼명
_HP_COL_CANDIDATES = ["홈페이지", "홈페이지주소", "웹사이트", "url", "homepage", "website"]


def _detect_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """DataFrame에서 후보 컬럼명 중 매칭되는 컬럼을 찾는다."""
    df_cols_lower = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        cand_lower = cand.lower()
        # 정확 매칭
        if cand_lower in df_cols_lower:
            return df_cols_lower[cand_lower]
        # 부분 매칭 (컬럼명에 후보가 포함된 경우)
        for col_lower, col_orig in df_cols_lower.items():
            if cand_lower in col_lower:
                return col_orig
    return None


def _parse_excel(file_path: Path) -> Tuple[pd.DataFrame, str, Optional[str], Optional[str]]:
    """
    엑셀 파일을 파싱하고 컬럼을 자동 감지한다.

    Returns:
        (df, name_col, addr_col, hp_col)
    """
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(file_path, dtype=str).fillna("")
    elif suffix == ".csv":
        try:
            df = pd.read_csv(file_path, dtype=str, encoding="utf-8").fillna("")
        except UnicodeDecodeError:
            df = pd.read_csv(file_path, dtype=str, encoding="cp949", encoding_errors="replace").fillna("")
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix} (xlsx, xls, csv만 지원)")

    if df.empty:
        raise ValueError("엑셀 파일이 비어 있습니다.")

    # 컬럼 감지
    name_col = _detect_column(df, _NAME_COL_CANDIDATES)
    if not name_col:
        raise ValueError(
            f"회사명 컬럼을 찾을 수 없습니다.\n"
            f"파일의 컬럼: {list(df.columns)}\n"
            f"다음 중 하나의 컬럼명을 사용해주세요: {_NAME_COL_CANDIDATES}"
        )

    addr_col = _detect_column(df, _ADDR_COL_CANDIDATES)
    hp_col = _detect_column(df, _HP_COL_CANDIDATES)

    log.info("컬럼 감지 결과:")
    log.info("  회사명: '%s'", name_col)
    log.info("  주소:   '%s'", addr_col or "(없음 — 검색 정확도 다소 낮아질 수 있음)")
    log.info("  홈페이지: '%s'", hp_col or "(없음)")

    return df, name_col, addr_col, hp_col


def _check_existing(
    companies: List[str],
    state: EnrichmentState,
) -> Tuple[Set[str], Dict[str, List[dict]]]:
    """
    이미 수집된 회사를 조회한다.

    Returns:
        (existing_norm_names, existing_results)
        - existing_norm_names: 이미 수집된 회사의 정규화된 이름 set
        - existing_results: 회사명(원본) → 기존 이메일 결과 리스트
    """
    all_emails = state.get_all_emails()
    existing_norm_names: Set[str] = set()
    existing_results: Dict[str, List[dict]] = {}

    # DB에 있는 모든 회사의 정규화명 → 원본 이메일 데이터 매핑
    norm_to_emails: Dict[str, List[dict]] = {}
    for em in all_emails:
        norm = _normalize_name(em.get("회사명", ""))
        if norm:
            norm_to_emails.setdefault(norm, []).append(em)

    # 입력 회사 목록과 매칭
    for company in companies:
        norm = _normalize_name(company)
        if norm and norm in norm_to_emails:
            existing_norm_names.add(norm)
            existing_results[company] = norm_to_emails[norm]

    return existing_norm_names, existing_results


def _style_sheet(ws, df: pd.DataFrame) -> None:
    """그레이스케일 엑셀 스타일링: 헤더 #D9D9D9, bold, freeze A2."""
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).fillna("").head(200)
        width = max([len(str(col))] + [min(len(str(v)), 60) for v in sample]) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(12, width), 60)

    ws.freeze_panes = "A2"


def run_enrich_excel(
    file_path: str,
    skip_existing: bool = False,
    limit: int = 0,
    force_recollect: bool = False,
) -> str:
    """
    사용자가 제공한 엑셀 파일의 회사 목록을 대상으로 이메일을 수집한다.

    Parameters
    ----------
    file_path : str
        입력 엑셀/CSV 파일 경로.
    skip_existing : bool
        True이면 이미 수집된 회사를 완전히 건너뛴다 (결과에도 미포함).
    limit : int
        0이면 전체, > 0이면 상위 N개만 처리.
    force_recollect : bool
        True이면 이미 수집된 회사도 재수집한다.

    Returns
    -------
    str
        출력 엑셀 파일 경로.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    # ── 1) 엑셀 파싱 & 컬럼 감지 ─────────────────────────────
    df, name_col, addr_col, hp_col = _parse_excel(path)

    if limit > 0:
        df = df.head(limit)

    total = len(df)
    log.info("입력 파일: %s (%d개 회사)", path.name, total)

    # ── 2) 중복 체크 ──────────────────────────────────────────
    state = EnrichmentState()
    companies = [str(row.get(name_col, "")).strip() for _, row in df.iterrows()]
    existing_norm_names, existing_results = _check_existing(companies, state)

    # 통계 출력
    new_count = sum(1 for c in companies if c and _normalize_name(c) not in existing_norm_names)
    dup_count = sum(1 for c in companies if c and _normalize_name(c) in existing_norm_names)
    empty_count = sum(1 for c in companies if not c.strip())

    log.info("=" * 50)
    log.info("  📊 분석 결과:")
    log.info("    전체: %d개 회사", total)
    log.info("    신규: %d개 (이메일 수집 예정)", new_count)
    log.info("    이미 수집: %d개%s", dup_count, " (스킵)" if skip_existing else " (기존 결과 사용)")
    if empty_count:
        log.info("    회사명 없음: %d개 (스킵)", empty_count)
    log.info("=" * 50)

    # ── 3) 수집 실행 ──────────────────────────────────────────
    out_rows: List[dict] = []
    stats = {STATUS_COLLECTED: 0, STATUS_NOT_FOUND: 0, STATUS_FAILED: 0, STATUS_SKIPPED: 0}
    company_no = 0

    for i, row in tqdm(list(df.iterrows()), total=len(df), desc="이메일 수집"):
        company = str(row.get(name_col, "")).strip()
        address = str(row.get(addr_col, "")).strip() if addr_col else ""
        homepage = str(row.get(hp_col, "")).strip() if hp_col else ""

        # 원본 행 데이터 보존
        base = row.to_dict()
        company_no += 1

        # (a) 회사명 없음
        if not company:
            if not skip_existing:
                out_rows.append({
                    **base,
                    "_No": company_no,
                    "이메일": "",
                    "이메일_출처_URL": "",
                    "이메일_방법": "",
                    "수집시각": "",
                    "상태": "회사명 없음",
                })
            stats[STATUS_SKIPPED] += 1
            continue

        norm = _normalize_name(company)

        # (b) 이미 수집된 회사
        if norm in existing_norm_names and not force_recollect:
            if skip_existing:
                tqdm.write(f"  ⏩ [{company_no}/{total}] {company} — 이미 수집됨 (스킵)")
                stats[STATUS_SKIPPED] += 1
                continue

            # 기존 결과를 출력에 포함
            existing = existing_results.get(company, [])
            if existing:
                for em in existing:
                    if em.get("상태") == STATUS_COLLECTED and em.get("이메일"):
                        out_rows.append({
                            **base,
                            "_No": company_no,
                            "이메일": em.get("이메일", ""),
                            "이메일_출처_URL": em.get("이메일_출처_URL", ""),
                            "이메일_방법": em.get("이메일_방법", ""),
                            "수집시각": em.get("조회시각", ""),
                            "상태": "기존 수집 결과",
                        })
                    else:
                        out_rows.append({
                            **base,
                            "_No": company_no,
                            "이메일": "",
                            "이메일_출처_URL": "",
                            "이메일_방법": "",
                            "수집시각": em.get("조회시각", ""),
                            "상태": f"기존 결과 ({em.get('상태', '')})",
                        })
            else:
                out_rows.append({
                    **base,
                    "_No": company_no,
                    "이메일": "",
                    "이메일_출처_URL": "",
                    "이메일_방법": "",
                    "수집시각": "",
                    "상태": "이미 수집됨 (결과 없음)",
                })

            tqdm.write(f"  ♻️  [{company_no}/{total}] {company} — 기존 결과 사용")
            stats[STATUS_SKIPPED] += 1
            continue

        # (c) 신규 수집
        tqdm.write(f"\n  🔍 [{company_no}/{total}] {company} 이메일 검색 중...")

        try:
            hits = search_emails_for_company(company, address, homepage_url=homepage)
        except Exception as e:
            log.warning("크롤링 실패 [%s]: %s", company, e)
            hits = []
            # 회사키 생성 (정규화 이름 기반)
            회사키 = f"excel_{norm}"
            state.upsert(회사키, company, 상태=STATUS_FAILED, 에러메시지=str(e))
            out_rows.append({
                **base,
                "_No": company_no,
                "이메일": "",
                "이메일_출처_URL": "",
                "이메일_방법": "",
                "수집시각": datetime.now().isoformat(timespec="seconds"),
                "상태": "수집 실패",
            })
            stats[STATUS_FAILED] += 1
            continue

        회사키 = f"excel_{norm}"

        if not hits:
            state.upsert(회사키, company, 상태=STATUS_NOT_FOUND)
            out_rows.append({
                **base,
                "_No": company_no,
                "이메일": "",
                "이메일_출처_URL": "",
                "이메일_방법": "",
                "수집시각": datetime.now().isoformat(timespec="seconds"),
                "상태": "이메일 미발견",
            })
            tqdm.write(f"    ❌ 이메일을 찾지 못했습니다.")
            stats[STATUS_NOT_FOUND] += 1
        else:
            for h in hits:
                state.upsert(
                    회사키, company,
                    이메일=h.email,
                    출처URL=h.source_url,
                    방법=h.method,
                    상태=STATUS_COLLECTED,
                )
                out_rows.append({
                    **base,
                    "_No": company_no,
                    "이메일": h.email,
                    "이메일_출처_URL": h.source_url,
                    "이메일_방법": h.method,
                    "수집시각": h.found_at,
                    "상태": "수집 완료",
                })
            tqdm.write(f"    ✅ {len(hits)}개 이메일 발견")
            stats[STATUS_COLLECTED] += 1

    state.close()

    # ── 4) 결과 엑셀 생성 ─────────────────────────────────────
    if not out_rows:
        log.warning("수집 결과가 없습니다.")
        return ""

    result_df = pd.DataFrame(out_rows)

    # No. 컬럼을 맨 앞으로, 결과 컬럼들을 맨 뒤로 정렬
    result_cols = ["이메일", "이메일_출처_URL", "이메일_방법", "수집시각", "상태"]
    original_cols = [c for c in result_df.columns if c not in result_cols and c != "_No"]
    ordered_cols = ["_No"] + original_cols + result_cols
    ordered_cols = [c for c in ordered_cols if c in result_df.columns]
    result_df = result_df[ordered_cols]
    result_df = result_df.rename(columns={"_No": "No."})

    # 출력 파일명: enriched_<원본파일명>_<날짜>.xlsx
    stem = path.stem
    out_path = OUTPUT_DIR / f"enriched_{stem}_{date.today().isoformat()}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 시트 1: 수집 결과 (전체)
        result_df.to_excel(writer, sheet_name="수집결과", index=False)
        ws = writer.book["수집결과"]
        _style_sheet(ws, result_df)

        # 시트 2: 요약
        summary_data = {
            "항목": ["전체 회사 수", "신규 수집", "이메일 발견", "이메일 미발견", "수집 실패", "기존 결과 사용/스킵", "회사명 없음"],
            "건수": [
                total,
                stats[STATUS_COLLECTED] + stats[STATUS_NOT_FOUND] + stats[STATUS_FAILED],
                stats[STATUS_COLLECTED],
                stats[STATUS_NOT_FOUND],
                stats[STATUS_FAILED],
                dup_count,
                empty_count,
            ],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="요약", index=False)
        ws_sum = writer.book["요약"]
        _style_sheet(ws_sum, summary_df)

    # ── 5) 최종 통계 출력 ─────────────────────────────────────
    log.info("=" * 50)
    log.info("  ✅ 수집 완료!")
    log.info("    이메일 발견: %d개 회사", stats[STATUS_COLLECTED])
    log.info("    이메일 미발견: %d개 회사", stats[STATUS_NOT_FOUND])
    log.info("    수집 실패: %d개 회사", stats[STATUS_FAILED])
    log.info("    스킵/기존 결과: %d개 회사", stats[STATUS_SKIPPED])
    total_searched = stats[STATUS_COLLECTED] + stats[STATUS_NOT_FOUND] + stats[STATUS_FAILED]
    if total_searched > 0:
        hit_rate = stats[STATUS_COLLECTED] / total_searched * 100
        log.info("    hit rate: %.1f%%", hit_rate)
    log.info("    결과 파일: %s", out_path)
    log.info("=" * 50)

    return str(out_path)

def run_enrich_excel_fast(file_path: str, limit: int = 0) -> str:
    import concurrent.futures
    import threading
    
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    df, name_col, addr_col, hp_col = _parse_excel(path)
    if limit > 0:
        df = df.head(limit)

    total = len(df)
    log.info("최적화/단일 추출 모드 시작: %s (%d개 회사)", path.name, total)

    out_rows = []
    n_with_email = 0
    save_counter = 0
    lock = threading.Lock()
    
    stem = path.stem
    out_path = OUTPUT_DIR / f"enriched_fast_{stem}_{date.today().isoformat()}.xlsx"

    def _process_row(idx, row):
        company = str(row.get(name_col, "")).strip()
        homepage = str(row.get(hp_col, "")).strip() if hp_col else ""
        base = row.to_dict()
        base["_No"] = idx + 1
        
        if not company:
            return {**base, "이메일": "", "이메일_출처_URL": "", "이메일_방법": "", "수집시각": "", "상태": "회사명 없음"}
            
        hit = fast_search_single_email(company, homepage)
        
        if hit:
            return {
                **base,
                "이메일": hit.email,
                "이메일_출처_URL": hit.source_url,
                "이메일_방법": hit.method,
                "수집시각": hit.found_at,
                "상태": "수집 완료"
            }
        else:
            return {
                **base,
                "이메일": "",
                "이메일_출처_URL": "",
                "이메일_방법": "",
                "수집시각": datetime.now().isoformat(timespec="seconds"),
                "상태": "이메일 미발견"
            }

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(_process_row, i, row): i for i, row in df.iterrows()}
        for future in tqdm(concurrent.futures.as_completed(futures), total=len(futures), desc="병렬 수집(Fast)"):
            try:
                res = future.result()
                with lock:
                    out_rows.append(res)
                    if res.get("이메일"):
                        n_with_email += 1
                    save_counter += 1
                    if save_counter % 100 == 0:
                        pd.DataFrame(out_rows).to_excel(out_path, index=False)
            except Exception as e:
                log.error(f"Task 처리 중 에러 발생: {e}")

    result_df = pd.DataFrame(out_rows)
    if "_No" in result_df.columns:
        result_df = result_df.sort_values("_No")
        
    result_cols = ["이메일", "이메일_출처_URL", "이메일_방법", "수집시각", "상태"]
    original_cols = [c for c in result_df.columns if c not in result_cols and c != "_No"]
    ordered_cols = ["_No"] + original_cols + result_cols
    ordered_cols = [c for c in ordered_cols if c in result_df.columns]
    result_df = result_df[ordered_cols]
    result_df = result_df.rename(columns={"_No": "No."})

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="수집결과", index=False)
        _style_sheet(writer.book["수집결과"], result_df)

    log.info("=" * 50)
    log.info("  ✅ 최적화 모드 수집 완료!")
    log.info("    전체 처리: %d건", total)
    log.info("    이메일 발견: %d건", n_with_email)
    log.info("    결과 파일: %s", out_path)
    log.info("=" * 50)

    return str(out_path)
